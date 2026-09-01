from __future__ import annotations

import csv
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook

from cassn.reporting.data_collection_summary import (
    MetadataDocument,
    WISubmissionTracker,
    build_data_collection_summary,
    dated_output_path,
    default_box_year_root,
    discover_box_year_metadata,
    load_wi_submission_tracker,
    render_data_collection_summary_workbook,
)


def _image(
    deployment_id: str,
    filename: str,
    *,
    site_name: str = "Sedgwick Reserve",
    box: str = "True",
    wi: str = "False",
    size: str = "1000",
) -> dict[str, str]:
    return {
        "deployment_id": deployment_id,
        "filename": filename,
        "site_name": site_name,
        "file_type": "image",
        "file_size_bytes": size,
        "is_uploaded_to_box": box,
        "is_submitted_to_wi": wi,
    }


def _audio(
    deployment_id: str,
    filename: str,
    device_type: str,
    seconds: str,
    *,
    site_name: str = "Sedgwick Reserve",
    box: str = "True",
    soundhub: str = "False",
    size: str = "2000",
) -> dict[str, str]:
    return {
        "deployment_id": deployment_id,
        "filename": filename,
        "site_name": site_name,
        "file_type": "audio",
        "device_type": device_type,
        "recording_duration_sec": seconds,
        "file_size_bytes": size,
        "is_uploaded_to_box": box,
        "is_submitted_to_soundhub": soundhub,
    }


def test_summary_uses_full_site_name_box_provenance_and_wi_tracker(tmp_path):
    wi_deployment = "UC_Sedgwick_plot1_ML_20260610"
    box_only_deployment = "UC_Sedgwick_plot2_ML_20260610"
    documents = [
        MetadataDocument(
            tmp_path / "image_file_metadata.csv",
            "image",
            (
                _image(wi_deployment, "one.jpg"),
                _image(box_only_deployment, "two.jpg"),
                _image(wi_deployment, "not-confirmed.jpg", box="False"),
            ),
        ),
        MetadataDocument(
            tmp_path / "audio_file_metadata.csv",
            "audio",
            (
                _audio("bird", "bird.wav", "BD", "120", soundhub="True"),
                _audio("bat", "bat.wav", "BT", "60"),
            ),
        ),
    ]
    tracker = WISubmissionTracker(
        tmp_path / "tracker.xlsx",
        {
            wi_deployment: "WI - Validation not started",
            box_only_deployment: "Box",
        },
    )

    result = build_data_collection_summary(2026, documents, wi_tracker=tracker)

    assert result.ok
    assert result.site_count == 1
    assert list(result.by_site) == ["Sedgwick Reserve"]
    assert result.totals.images_captured == 2
    assert result.totals.images_classified_ai == 1
    assert result.totals.bird_minutes_recorded == Decimal("2")
    assert result.totals.bird_minutes_classified_ai == Decimal("2")
    assert result.totals.bat_minutes_recorded == Decimal("1")
    assert result.totals.data_bytes == 6000
    assert result.quality_counts["unconfirmed_box_upload_rows"] == 1
    assert result.quality_counts["wi_tracker_metadata_backfill_deployments"] == 1
    assert result.quality_counts["wi_tracker_metadata_conflicts"] == 0


def test_missing_duration_does_not_use_schedule_duration(tmp_path):
    row = _audio("bird", "bird.wav", "BD", "")
    row["duration"] = "12:00"
    result = build_data_collection_summary(
        2026,
        [MetadataDocument(tmp_path / "audio_file_metadata.csv", "audio", (row,))],
    )

    assert result.totals.bird_minutes_recorded == 0
    assert result.quality_counts["missing_audio_duration_rows"] == 1


def test_duplicate_key_with_same_hash_is_counted_once(tmp_path):
    first = _audio("bird", "bird.wav", "BD", "120")
    first["file_hash_sha256"] = "abc123"
    second = dict(first)
    second["legacy_extra_column"] = "historical copy"
    result = build_data_collection_summary(
        2026,
        [MetadataDocument(tmp_path / "audio_file_metadata.csv", "audio", (first, second))],
    )

    assert result.ok
    assert result.totals.bird_minutes_recorded == 2
    assert result.quality_counts["exact_duplicate_rows"] == 1


def test_duplicate_blank_site_name_is_only_backfilled_once(tmp_path):
    first = _audio("bird", "bird.wav", "BD", "120", site_name="")
    first["file_hash_sha256"] = "abc123"
    second = dict(first)
    result = build_data_collection_summary(
        2026,
        [
            MetadataDocument(
                tmp_path / "audio_file_metadata.csv",
                "audio",
                (first, second),
                "Quail Ridge Reserve",
            )
        ],
    )

    assert result.ok
    assert result.quality_counts["exact_duplicate_rows"] == 1
    assert result.quality_counts["site_name_from_box_folder_rows"] == 1


def test_discovers_only_event_root_metadata(tmp_path):
    event = tmp_path / "Reserve" / "Event"
    raw = event / "raw_data" / "p1_ML"
    raw.mkdir(parents=True)
    fields = list(_image("deployment", "one.jpg"))
    with (event / "image_file_metadata.csv").open("w", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fields)
        writer.writeheader()
        writer.writerow(_image("deployment", "one.jpg"))
    (raw / "image_file_metadata.csv").write_text("should,not,be,read\n")

    documents = discover_box_year_metadata(tmp_path)

    assert len(documents) == 1
    assert documents[0].path == event / "image_file_metadata.csv"
    assert documents[0].site_name_hint == "Reserve"


def test_blank_legacy_site_name_uses_full_box_reserve_folder_name(tmp_path):
    row = _image("deployment", "one.jpg", site_name="")
    result = build_data_collection_summary(
        2026,
        [
            MetadataDocument(
                tmp_path / "image_file_metadata.csv",
                "image",
                (row,),
                "Angelo Coast Range Reserve",
            )
        ],
    )

    assert result.ok
    assert list(result.by_site) == ["Angelo Coast Range Reserve"]
    assert result.quality_counts["site_name_from_box_folder_rows"] == 1


def test_loads_exported_wi_tracker_workbook(tmp_path):
    path = tmp_path / "tracker.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Deployment Tracker"
    sheet.append(["Project", "Deployment", "Status"])
    sheet.append(["UCNRS Mammal", "deployment-1", "WI - Partial Validation"])
    sheet.append(["UCNRS Mammal", "deployment-2", "Box"])
    workbook.save(path)
    workbook.close()

    tracker = load_wi_submission_tracker(path)

    assert tracker.submitted("deployment-1") is True
    assert tracker.submitted("deployment-2") is False
    assert tracker.submitted("missing") is None


def test_renders_only_the_three_approved_sheets_in_requested_style(tmp_path):
    result = build_data_collection_summary(
        2026,
        [
            MetadataDocument(
                tmp_path / "image_file_metadata.csv",
                "image",
                (_image("deployment", "one.jpg", wi="True"),),
            )
        ],
    )
    path = tmp_path / "summary.xlsx"

    render_data_collection_summary_workbook(
        result,
        path,
        generated_at=datetime.fromisoformat("2026-08-31T12:00:00-07:00"),
    )

    workbook = load_workbook(path, data_only=True)
    try:
        assert workbook.sheetnames == ["Summary", "By Site", "Data Quality"]
        summary_sheet = workbook["Summary"]
        by_site_sheet = workbook["By Site"]
        quality_sheet = workbook["Data Quality"]

        assert summary_sheet["A1"].value == "CA-SSN 2026 Data Collection Summary"
        assert summary_sheet["A2"].value == (
            "Box-uploaded media stored in CASSN/data/2026 as of 2026-08-31."
        )
        assert summary_sheet["A5"].value == "Sites with collected data"
        assert summary_sheet["B5"].value == 1
        assert "; no contractor-classification total yet" not in summary_sheet["D10"].value

        assert by_site_sheet["A1"].value == "CA-SSN 2026 Data Collection by Site"
        assert by_site_sheet["A2"].value is None
        assert by_site_sheet["A3"].value is None
        assert by_site_sheet["A4"].value == "Site Name"
        assert by_site_sheet["A5"].value == "Sedgwick Reserve"

        for sheet, last_column in (
            (summary_sheet, 4),
            (by_site_sheet, 7),
            (quality_sheet, 4),
        ):
            header_cell = sheet.cell(4, 1)
            assert header_cell.fill.fill_type == "solid"
            assert header_cell.fill.fgColor.rgb == "FF000000"
            assert header_cell.font.bold
            assert header_cell.font.color.rgb.endswith("FFFFFF")
            for row in range(4, sheet.max_row + 1):
                for column in range(1, last_column + 1):
                    cell = sheet.cell(row, column)
                    assert all(
                        edge.style == "thin"
                        for edge in (
                            cell.border.left,
                            cell.border.right,
                            cell.border.top,
                            cell.border.bottom,
                        )
                    )

            outside_cell = sheet.cell(4, last_column + 1)
            assert outside_cell.fill.fill_type is None
            assert all(
                edge.style is None
                for edge in (
                    outside_cell.border.left,
                    outside_cell.border.right,
                    outside_cell.border.top,
                    outside_cell.border.bottom,
                )
            )

        assert summary_sheet["A5"].fill.fgColor.rgb.endswith("D9D9D9")
        assert summary_sheet["A6"].fill.fgColor.rgb.endswith("FFFFFF")
    finally:
        workbook.close()


def test_dated_output_path_uses_generation_date_folder(tmp_path):
    path = dated_output_path(tmp_path, 2026, datetime(2026, 8, 31).date())
    assert path == tmp_path / "2026-08-31" / "CA-SSN_data_collection_summary_2026.xlsx"


def test_default_box_year_root_uses_renamed_data_folder():
    assert default_box_year_root(2026).parts[-3:] == ("CASSN", "data", "2026")
