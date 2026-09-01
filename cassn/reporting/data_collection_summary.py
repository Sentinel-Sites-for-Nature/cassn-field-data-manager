"""Summarize confirmed Box-uploaded field data for one Box year folder.

The reporting boundary is deliberately the selected ``data/<year>``
folder, not the calendar year parsed from each recording timestamp.  The
scanner reads only the two small metadata CSVs at each deployment-event root;
it never enumerates or opens raw media.

Wildlife Insights is submitted manually, so the optional WI tracker is a
second authority keyed by ``deployment_id``.  A tracker status beginning
``WI -`` proves that the deployment reached Wildlife Insights.  ``Box`` means
it has not.  For deployments absent from the tracker, the durable metadata
field ``is_submitted_to_wi`` remains the fallback.
"""

from __future__ import annotations

import csv
import os
import tempfile
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from cassn.config import PACIFIC_TZ


TRUE_VALUES = frozenset({"1", "true", "yes", "y"})
FALSE_VALUES = frozenset({"", "0", "false", "no", "n"})
SUMMARY_FILENAME_TEMPLATE = "CA-SSN_data_collection_summary_{year}.xlsx"
BOX_REPORTS_FOLDER_ID = "413722950229"


def default_box_year_root(year: int) -> Path:
    """Return the normal Box Drive data root for ``year``."""
    return (
        Path.home()
        / "Library"
        / "CloudStorage"
        / "Box-Box"
        / "CASSN"
        / "data"
        / str(year)
    )


def default_box_reports_root() -> Path:
    """Return the Box Drive folder corresponding to Box id 413722950229."""
    return (
        Path.home()
        / "Library"
        / "CloudStorage"
        / "Box-Box"
        / "CASSN"
        / "data_collection_summary_stats"
    )


@dataclass(frozen=True)
class MetadataDocument:
    """One filed metadata CSV and its decoded rows."""

    path: Path
    kind: str
    rows: tuple[dict[str, str], ...]
    site_name_hint: str = ""


@dataclass(frozen=True)
class WISubmissionTracker:
    """Deployment-level WI state loaded from John's validation tracker."""

    source: Path
    status_by_deployment: dict[str, str]
    warnings: tuple[str, ...] = ()

    def submitted(self, deployment_id: str) -> bool | None:
        status = self.status_by_deployment.get(deployment_id)
        if status is None:
            return None
        normalized = status.strip().lower()
        if normalized.startswith("wi -"):
            return True
        if normalized == "box":
            return False
        return None


@dataclass
class SiteTotals:
    site_name: str
    images_captured: int = 0
    images_classified_ai: int = 0
    bird_seconds_recorded: Decimal = Decimal(0)
    bird_seconds_classified_ai: Decimal = Decimal(0)
    bat_seconds_recorded: Decimal = Decimal(0)
    data_bytes: int = 0

    @property
    def bird_minutes_recorded(self) -> Decimal:
        return self.bird_seconds_recorded / Decimal(60)

    @property
    def bird_minutes_classified_ai(self) -> Decimal:
        return self.bird_seconds_classified_ai / Decimal(60)

    @property
    def bat_minutes_recorded(self) -> Decimal:
        return self.bat_seconds_recorded / Decimal(60)

    @property
    def data_gb(self) -> Decimal:
        return Decimal(self.data_bytes) / Decimal(1_000_000_000)


@dataclass
class DataCollectionSummary:
    year: int
    totals: SiteTotals = field(default_factory=lambda: SiteTotals("All sites"))
    by_site: dict[str, SiteTotals] = field(default_factory=dict)
    quality_counts: Counter = field(default_factory=Counter)
    quality_details: dict[str, str] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    metadata_files_scanned: int = 0
    wi_tracker_source: str = ""

    @property
    def site_count(self) -> int:
        return len(self.by_site)

    @property
    def ok(self) -> bool:
        return not self.errors


def _read_csv(path: Path) -> tuple[dict[str, str], ...]:
    with path.open(newline="", encoding="utf-8-sig") as stream:
        reader = csv.DictReader(stream)
        if reader.fieldnames is None:
            raise ValueError("CSV has no header")
        rows = tuple(
            {str(key): str(value or "").strip() for key, value in row.items() if key is not None}
            for row in reader
        )
    return rows


def discover_box_year_metadata(box_year_root: Path) -> list[MetadataDocument]:
    """Read event-root metadata without descending into raw-data folders."""
    root = Path(box_year_root).expanduser().resolve()
    if not root.is_dir():
        raise FileNotFoundError(f"Box year root does not exist: {root}")

    documents: list[MetadataDocument] = []
    for reserve in sorted(path for path in root.iterdir() if path.is_dir()):
        for event in sorted(path for path in reserve.iterdir() if path.is_dir()):
            for filename, kind in (
                ("image_file_metadata.csv", "image"),
                ("audio_file_metadata.csv", "audio"),
            ):
                path = event / filename
                if not path.is_file():
                    continue
                try:
                    rows = _read_csv(path)
                except Exception as exc:
                    raise ValueError(f"Could not read {path}: {exc}") from exc
                documents.append(MetadataDocument(path, kind, rows, reserve.name))
    return documents


def _tracker_from_rows(source: Path, rows: list[dict[str, str]]) -> WISubmissionTracker:
    statuses: dict[str, str] = {}
    warnings: list[str] = []
    for row_number, row in enumerate(rows, start=2):
        deployment_id = str(row.get("Deployment") or "").strip()
        status = str(row.get("Status") or "").strip()
        if not deployment_id:
            continue
        if deployment_id in statuses and statuses[deployment_id] != status:
            raise ValueError(
                f"{source}: conflicting WI statuses for {deployment_id!r}"
            )
        statuses[deployment_id] = status
        normalized = status.lower()
        if normalized != "box" and not normalized.startswith("wi -"):
            warnings.append(
                f"{source.name} row {row_number}: unrecognized WI status {status!r}"
            )
    if not statuses:
        raise ValueError(f"{source}: no Deployment rows found")
    return WISubmissionTracker(source, statuses, tuple(warnings))


def load_wi_submission_tracker(path: Path) -> WISubmissionTracker:
    """Load the Deployment Tracker tab from a CSV or exported XLSX workbook."""
    source = Path(path).expanduser().resolve()
    if not source.is_file():
        raise FileNotFoundError(f"WI tracker does not exist: {source}")
    if source.suffix.lower() == ".csv":
        return _tracker_from_rows(source, list(_read_csv(source)))
    if source.suffix.lower() != ".xlsx":
        raise ValueError("WI tracker must be a .csv or .xlsx file")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is required by the app
        raise RuntimeError("openpyxl is required to read the WI tracker") from exc

    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        if "Deployment Tracker" not in workbook.sheetnames:
            raise ValueError(f"{source}: no 'Deployment Tracker' sheet")
        sheet = workbook["Deployment Tracker"]
        values = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(values)]
        rows = [
            {headers[index]: str(value or "").strip() for index, value in enumerate(row)}
            for row in values
        ]
    finally:
        workbook.close()
    return _tracker_from_rows(source, rows)


def _boolean_state(value: object) -> bool | None:
    normalized = str(value or "").strip().lower()
    if normalized in TRUE_VALUES:
        return True
    if normalized in FALSE_VALUES:
        return False
    return None


def _duration_seconds(value: object) -> Decimal | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        seconds = Decimal(text)
    except InvalidOperation:
        return None
    return seconds if seconds >= 0 else None


def _file_size(value: object) -> int | None:
    try:
        size = int(str(value or "").strip())
    except (TypeError, ValueError):
        return None
    return size if size >= 0 else None


def _same_media_bytes(first: dict[str, str], second: dict[str, str]) -> bool:
    """Whether duplicate metadata keys describe the same filed media bytes."""
    for hash_field in ("file_hash_sha256", "file_hash_sha1"):
        left = str(first.get(hash_field) or "").strip().lower()
        right = str(second.get(hash_field) or "").strip().lower()
        if left and right:
            return left == right
    left_size = _file_size(first.get("file_size_bytes"))
    right_size = _file_size(second.get("file_size_bytes"))
    return left_size is not None and left_size == right_size


def _add_detail(summary: DataCollectionSummary, key: str, detail: str) -> None:
    if detail and key not in summary.quality_details:
        summary.quality_details[key] = detail


def build_data_collection_summary(
    year: int,
    documents: list[MetadataDocument],
    *,
    wi_tracker: WISubmissionTracker | None = None,
) -> DataCollectionSummary:
    """Aggregate confirmed Box media into overall and full-site-name totals."""
    summary = DataCollectionSummary(year=year, metadata_files_scanned=len(documents))
    summary.quality_counts["metadata_files_scanned"] = len(documents)
    if wi_tracker is None:
        summary.warnings.append(
            "No WI tracker supplied; image classification uses metadata provenance only."
        )
        summary.quality_counts["wi_tracker_not_supplied"] = 1
    else:
        summary.wi_tracker_source = str(wi_tracker.source)
        summary.warnings.extend(wi_tracker.warnings)

    seen: dict[tuple[str, str], dict[str, str]] = {}
    wi_missing_deployments: set[str] = set()
    wi_tracker_backfills: set[str] = set()
    wi_conflicts: set[str] = set()
    tracker_deployments_seen: set[str] = set()

    for document in documents:
        for row_number, row in enumerate(document.rows, start=2):
            file_type = row.get("file_type", "").strip().lower()
            if document.kind == "image" and file_type != "image":
                continue
            if document.kind == "audio" and file_type != "audio":
                continue

            box_state = _boolean_state(row.get("is_uploaded_to_box"))
            if box_state is not True:
                summary.quality_counts["unconfirmed_box_upload_rows"] += 1
                if box_state is None:
                    summary.quality_counts["invalid_box_provenance_rows"] += 1
                continue

            site_name = row.get("site_name", "").strip()
            deployment_id = row.get("deployment_id", "").strip()
            filename = row.get("filename", "").strip()
            if not deployment_id or not filename:
                summary.errors.append(
                    f"{document.path} row {row_number}: blank deployment_id or filename"
                )
                summary.quality_counts["missing_media_key_rows"] += 1
                continue

            key = (deployment_id, filename)
            prior = seen.get(key)
            if prior is not None:
                if _same_media_bytes(prior, row):
                    summary.quality_counts["exact_duplicate_rows"] += 1
                    _add_detail(
                        summary,
                        "exact_duplicate_rows",
                        "Same deployment_id/filename and matching hash or size; counted once.",
                    )
                    continue
                summary.errors.append(
                    f"Conflicting metadata rows for {deployment_id}/{filename}"
                )
                summary.quality_counts["conflicting_duplicate_rows"] += 1
                continue
            seen[key] = row

            if not site_name and document.site_name_hint:
                site_name = document.site_name_hint
                summary.quality_counts["site_name_from_box_folder_rows"] += 1
                _add_detail(
                    summary,
                    "site_name_from_box_folder_rows",
                    "Blank legacy site_name filled from the enclosing Box reserve folder.",
                )
            if not site_name:
                summary.errors.append(f"{document.path} row {row_number}: blank site_name")
                summary.quality_counts["missing_site_name_rows"] += 1
                continue

            site = summary.by_site.setdefault(site_name, SiteTotals(site_name))
            size = _file_size(row.get("file_size_bytes"))
            if size is None:
                summary.quality_counts["invalid_file_size_rows"] += 1
                _add_detail(
                    summary,
                    "invalid_file_size_rows",
                    "Media row counted, but excluded from the total data volume.",
                )
            else:
                site.data_bytes += size
                summary.totals.data_bytes += size

            summary.quality_counts["confirmed_box_uploaded_media_rows"] += 1
            if document.kind == "image":
                site.images_captured += 1
                summary.totals.images_captured += 1
                metadata_wi = _boolean_state(row.get("is_submitted_to_wi")) is True
                submitted = metadata_wi
                if wi_tracker is not None:
                    tracker_state = wi_tracker.submitted(deployment_id)
                    if tracker_state is None:
                        wi_missing_deployments.add(deployment_id)
                    else:
                        tracker_deployments_seen.add(deployment_id)
                        submitted = tracker_state
                        if tracker_state and not metadata_wi:
                            wi_tracker_backfills.add(deployment_id)
                        if not tracker_state and metadata_wi:
                            wi_conflicts.add(deployment_id)
                if submitted:
                    site.images_classified_ai += 1
                    summary.totals.images_classified_ai += 1
                continue

            device_type = row.get("device_type", "").strip().upper()
            seconds = _duration_seconds(row.get("recording_duration_sec"))
            if seconds is None:
                summary.quality_counts["missing_audio_duration_rows"] += 1
                _add_detail(
                    summary,
                    "missing_audio_duration_rows",
                    "Recording counted as filed media, but omitted from minute totals.",
                )
                continue
            if device_type == "BD":
                site.bird_seconds_recorded += seconds
                summary.totals.bird_seconds_recorded += seconds
                if _boolean_state(row.get("is_submitted_to_soundhub")) is True:
                    site.bird_seconds_classified_ai += seconds
                    summary.totals.bird_seconds_classified_ai += seconds
            elif device_type == "BT":
                site.bat_seconds_recorded += seconds
                summary.totals.bat_seconds_recorded += seconds
            else:
                summary.quality_counts["unsupported_audio_device_rows"] += 1

    if wi_tracker is not None:
        summary.quality_counts["wi_tracker_deployments_used"] = len(tracker_deployments_seen)
        summary.quality_counts["wi_tracker_missing_deployments"] = len(wi_missing_deployments)
        summary.quality_counts["wi_tracker_metadata_backfill_deployments"] = len(
            wi_tracker_backfills
        )
        summary.quality_counts["wi_tracker_metadata_conflicts"] = len(wi_conflicts)
        _add_detail(
            summary,
            "wi_tracker_metadata_backfill_deployments",
            "Tracker proves WI upload even though Box image metadata is not stamped.",
        )
        if wi_missing_deployments:
            _add_detail(
                summary,
                "wi_tracker_missing_deployments",
                "Box image deployment is not listed in the supplied WI tracker; metadata fallback used.",
            )
        if wi_conflicts:
            summary.errors.append(
                "WI tracker says Box-only while metadata says submitted for: "
                + ", ".join(sorted(wi_conflicts))
            )

    if not documents:
        summary.errors.append("No image_file_metadata.csv or audio_file_metadata.csv files found")
    return summary


def _generated_at(value: datetime | None = None) -> datetime:
    moment = value or datetime.now(PACIFIC_TZ)
    return moment if moment.tzinfo else moment.replace(tzinfo=PACIFIC_TZ)


def render_data_collection_summary_workbook(
    summary: DataCollectionSummary,
    output_path: Path,
    *,
    generated_at: datetime | None = None,
) -> Path:
    """Write the approved three-sheet XLSX report atomically."""
    if not summary.ok:
        raise ValueError("Cannot publish a summary with validation errors")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - dependency is required by the app
        raise RuntimeError("openpyxl is required to create .xlsx reports") from exc

    output = Path(output_path).expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    moment = _generated_at(generated_at)

    # Use explicit opaque ARGB values. Desktop Excel treats six-digit black as
    # black, but Box Preview can interpret its omitted/zero alpha as no fill,
    # leaving the white header text invisible on a white background.
    black = "FF000000"
    white = "FFFFFFFF"
    gray = "FF666666"
    light_gray = "FFD9D9D9"
    medium_gray = "FFBFBFBF"
    dark_gray = "FF7F7F7F"
    thin_black = Side(style="thin", color=black)
    table_border = Border(
        left=thin_black,
        right=thin_black,
        top=thin_black,
        bottom=thin_black,
    )

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    by_site_sheet = workbook.create_sheet("By Site")
    quality_sheet = workbook.create_sheet("Data Quality")

    def title(sheet, text: str, end_column: int) -> None:
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
        cell = sheet.cell(1, 1, text)
        cell.font = Font(name="Aptos Display", size=18, bold=True, color=black)
        cell.alignment = Alignment(vertical="center")
        sheet.row_dimensions[1].height = 30
        sheet.sheet_view.showGridLines = False

    def header(sheet, row: int, columns: int) -> None:
        for cell in sheet[row][:columns]:
            cell.font = Font(name="Aptos", bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=black)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = table_border
        sheet.row_dimensions[row].height = 30

    def body(sheet, first_row: int, last_row: int, columns: int) -> None:
        for row in range(first_row, last_row + 1):
            fill = PatternFill(
                "solid", fgColor=light_gray if (row - first_row) % 2 == 0 else white
            )
            for cell in sheet[row][:columns]:
                cell.fill = fill
                cell.border = table_border

    title(summary_sheet, f"CA-SSN {summary.year} Data Collection Summary", 4)
    summary_sheet["A2"] = (
        f"Box-uploaded media stored in CASSN/data/{summary.year} as of "
        f"{moment.strftime('%Y-%m-%d')}."
    )
    summary_sheet.merge_cells("A2:D2")
    summary_sheet["A2"].font = Font(name="Aptos", italic=True, color=gray)
    summary_sheet["A2"].alignment = Alignment(wrap_text=True)
    summary_sheet.append([])
    summary_sheet.append(["Metric", "Value", "Unit", "Definition"])
    metrics = [
        (
            "Sites with collected data",
            summary.site_count,
            "sites",
            "Distinct full site_name values represented by confirmed Box-uploaded media.",
        ),
        ("Images captured", summary.totals.images_captured, "images", "Confirmed Box image rows."),
        (
            "Images classified with AI",
            summary.totals.images_classified_ai,
            "images",
            "Images in WI-submitted deployments; WI classification follows upload automatically.",
        ),
        (
            "Bird audio recorded",
            float(summary.totals.bird_minutes_recorded),
            "minutes",
            "Exact BD recording_duration_sec total.",
        ),
        (
            "Bird audio classified with AI",
            float(summary.totals.bird_minutes_classified_ai),
            "minutes",
            "BD minutes submitted to SoundHub; classification follows submission.",
        ),
        (
            "Bat audio recorded",
            float(summary.totals.bat_minutes_recorded),
            "minutes",
            "Exact BT recording_duration_sec total.",
        ),
        (
            "Total data volume",
            float(summary.totals.data_gb),
            "GB",
            "Sum of media file_size_bytes using decimal gigabytes.",
        ),
    ]
    for row in metrics:
        summary_sheet.append(row)
    header(summary_sheet, 4, 4)
    body(summary_sheet, 5, 4 + len(metrics), 4)
    for row in range(5, 5 + len(metrics)):
        summary_sheet.cell(row, 1).font = Font(name="Aptos", bold=True)
        summary_sheet.cell(row, 2).alignment = Alignment(horizontal="right")
        summary_sheet.cell(row, 4).alignment = Alignment(wrap_text=True)
    summary_sheet["B5"].number_format = "#,##0"
    summary_sheet["B6"].number_format = "#,##0"
    summary_sheet["B7"].number_format = "#,##0"
    for row in (8, 9, 10):
        summary_sheet.cell(row, 2).number_format = "#,##0.0"
    summary_sheet["B11"].number_format = "#,##0.00"
    summary_sheet.freeze_panes = "A5"
    summary_sheet.auto_filter.ref = f"A4:D{4 + len(metrics)}"
    for column, width in enumerate((34, 18, 14, 72), start=1):
        summary_sheet.column_dimensions[get_column_letter(column)].width = width

    title(by_site_sheet, f"CA-SSN {summary.year} Data Collection by Site", 7)
    by_site_sheet.append([])
    by_site_sheet.append([])
    by_site_sheet.append(
        [
            "Site Name",
            "Images Captured",
            "Images Classified with AI",
            "Bird Minutes Recorded",
            "Bird Minutes Classified with AI",
            "Bat Minutes Recorded",
            "Data Volume (GB)",
        ]
    )
    for site_name in sorted(summary.by_site):
        site = summary.by_site[site_name]
        by_site_sheet.append(
            [
                site.site_name,
                site.images_captured,
                site.images_classified_ai,
                float(site.bird_minutes_recorded),
                float(site.bird_minutes_classified_ai),
                float(site.bat_minutes_recorded),
                float(site.data_gb),
            ]
        )
    header(by_site_sheet, 4, 7)
    last_site_row = max(4, 4 + len(summary.by_site))
    if summary.by_site:
        body(by_site_sheet, 5, last_site_row, 7)
    for row in range(5, last_site_row + 1):
        by_site_sheet.cell(row, 1).font = Font(name="Aptos", bold=True)
        for column in (2, 3):
            by_site_sheet.cell(row, column).number_format = "#,##0"
        for column in (4, 5, 6):
            by_site_sheet.cell(row, column).number_format = "#,##0.0"
        by_site_sheet.cell(row, 7).number_format = "#,##0.00"
    by_site_sheet.freeze_panes = "B5"
    if summary.by_site:
        by_site_sheet.auto_filter.ref = f"A4:G{last_site_row}"
    widths = (55, 18, 25, 21, 29, 20, 18)
    for column, width in enumerate(widths, start=1):
        by_site_sheet.column_dimensions[get_column_letter(column)].width = width

    title(quality_sheet, f"CA-SSN {summary.year} Data Quality Checks", 4)
    quality_sheet["A2"] = (
        "These checks explain exclusions or incomplete source metadata; they are not additional program statistics."
    )
    quality_sheet.merge_cells("A2:D2")
    quality_sheet["A2"].font = Font(name="Aptos", italic=True, color=gray)
    quality_sheet["A2"].alignment = Alignment(wrap_text=True)
    quality_sheet.append([])
    quality_sheet.append(["Check", "Status", "Count", "Detail"])

    quality_order = [
        ("metadata_files_scanned", "Metadata files scanned", "info"),
        ("confirmed_box_uploaded_media_rows", "Confirmed Box-uploaded media rows", "info"),
        ("unconfirmed_box_upload_rows", "Unconfirmed Box-upload rows excluded", "warning"),
        ("invalid_box_provenance_rows", "Invalid Box provenance values", "warning"),
        ("missing_audio_duration_rows", "Audio rows missing exact duration", "warning"),
        ("invalid_file_size_rows", "Media rows missing valid file size", "warning"),
        ("exact_duplicate_rows", "Exact duplicate metadata rows ignored", "warning"),
        ("conflicting_duplicate_rows", "Conflicting duplicate metadata rows", "error"),
        (
            "site_name_from_box_folder_rows",
            "Rows using full site name from Box reserve folder",
            "info",
        ),
        ("missing_site_name_rows", "Rows missing full site_name", "error"),
        ("missing_media_key_rows", "Rows missing deployment/file identity", "error"),
        ("wi_tracker_not_supplied", "WI tracker not supplied", "warning"),
        ("wi_tracker_deployments_used", "WI tracker deployments used", "info"),
        ("wi_tracker_missing_deployments", "Image deployments absent from WI tracker", "warning"),
        (
            "wi_tracker_metadata_backfill_deployments",
            "WI tracker submissions not stamped in metadata",
            "warning",
        ),
        ("wi_tracker_metadata_conflicts", "WI tracker/metadata conflicts", "error"),
    ]
    for key, label, severity in quality_order:
        count = int(summary.quality_counts.get(key, 0))
        if severity == "info":
            status = "INFO"
        elif count:
            status = "ERROR" if severity == "error" else "WARNING"
        else:
            status = "PASS"
        detail = summary.quality_details.get(key, "")
        quality_sheet.append([label, status, count, detail])
    for warning in summary.warnings:
        quality_sheet.append(["Generation warning", "WARNING", 1, warning])
    header(quality_sheet, 4, 4)
    last_quality_row = quality_sheet.max_row
    body(quality_sheet, 5, last_quality_row, 4)
    for row in range(5, last_quality_row + 1):
        status = quality_sheet.cell(row, 2).value
        fill_color = {
            "PASS": white,
            "WARNING": medium_gray,
            "ERROR": dark_gray,
            "INFO": light_gray,
        }.get(status, white)
        quality_sheet.cell(row, 2).fill = PatternFill("solid", fgColor=fill_color)
        quality_sheet.cell(row, 2).font = Font(
            name="Aptos", bold=True, color=white if status == "ERROR" else black
        )
        quality_sheet.cell(row, 2).alignment = Alignment(horizontal="center")
        quality_sheet.cell(row, 3).number_format = "#,##0"
        quality_sheet.cell(row, 4).alignment = Alignment(wrap_text=True)
    quality_sheet.freeze_panes = "A5"
    quality_sheet.auto_filter.ref = f"A4:D{last_quality_row}"
    for column, width in enumerate((44, 13, 12, 80), start=1):
        quality_sheet.column_dimensions[get_column_letter(column)].width = width

    for sheet in workbook.worksheets:
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.outlinePr.summaryBelow = True

    fd, tmp_name = tempfile.mkstemp(
        prefix=f".{output.name}.", suffix=".tmp.xlsx", dir=output.parent
    )
    os.close(fd)
    tmp_path = Path(tmp_name)
    try:
        workbook.save(tmp_path)
        os.replace(tmp_path, output)
    finally:
        workbook.close()
        if tmp_path.exists():
            tmp_path.unlink()
    return output


def dated_output_path(output_root: Path, year: int, generation_date: date) -> Path:
    """Return ``<root>/<date>/CA-SSN_data_collection_summary_<year>.xlsx``."""
    return (
        Path(output_root).expanduser()
        / generation_date.isoformat()
        / SUMMARY_FILENAME_TEMPLATE.format(year=year)
    )
